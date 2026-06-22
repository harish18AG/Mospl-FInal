"""
load_report_generator.py - Generates a styled Excel report for E2E Load Tests.
Fits all specific column and summary requirements.
"""
import os
import datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import PieChart, Reference
from openpyxl.chart.label import DataLabelList

OUTPUT_DIR = os.path.join(os.path.dirname(__file__), "..", "reports")

def _cell_fill(hex_color: str) -> PatternFill:
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")

def _thin_border() -> Border:
    s = Side(style="thin", color="D3D3D3")
    return Border(left=s, right=s, top=s, bottom=s)

def generate_load_test_report(test_results: list):
    """
    Generate an Excel report containing:
    - Executive Summary Sheet (with Summary Section + Chart)
    - Detailed Results Sheet (with 300 test cases)
    """
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    timestamp = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    filename = os.path.join(OUTPUT_DIR, f"Mospl_Load_Test_Report_{timestamp}.xlsx")

    # Convert results list to DataFrame
    df = pd.DataFrame(test_results)

    # 1. Compute summary statistics
    total = len(df)
    passed = (df["Status"] == "PASS").sum()
    failed = (df["Status"] == "FAIL").sum()
    pass_percentage = round((passed / total) * 100, 2) if total > 0 else 0.0
    
    # Calculate performance score: (Threshold / Measured Value) * 100 capped at 100 for each test case
    # If measured value is 0, score is 100.
    perf_scores = []
    for _, row in df.iterrows():
        try:
            measured = float(row["Measured Value"].split()[0])
            threshold = float(row["Threshold"].split()[0])
            if measured <= 0:
                score = 100.0
            else:
                score = min(100.0, (threshold / measured) * 100.0)
            perf_scores.append(score)
        except Exception:
            perf_scores.append(100.0 if row["Status"] == "PASS" else 50.0)
            
    avg_perf_score = round(sum(perf_scores) / len(perf_scores), 2) if perf_scores else 100.0

    # 2. Write sheets using pandas ExcelWriter
    with pd.ExcelWriter(filename, engine="openpyxl") as writer:
        # Sheet 1: Executive Summary
        summary_data = {
            "Summary Metric": [
                "Total Test Cases",
                "Passed",
                "Failed",
                "Pass Percentage",
                "Average Performance Score"
            ],
            "Value": [
                total,
                passed,
                failed,
                f"{pass_percentage}%",
                f"{avg_perf_score} / 100"
            ]
        }
        pd.DataFrame(summary_data).to_excel(writer, sheet_name="Executive Summary", index=False, startrow=2)

        # Sheet 2: Load Test Details
        details_df = df[[
            "Test Case",
            "Category",
            "Measured Value",
            "Threshold",
            "Result",
            "Status"
        ]].copy()
        details_df.to_excel(writer, sheet_name="Load Test Details", index=False, startrow=2)

    # 3. Open workbook with openpyxl to apply rich styling
    wb = load_workbook(filename)

    # ─── Style "Executive Summary" ──────────────────────────────────────────
    ws_sum = wb["Executive Summary"]
    # Title Block
    ws_sum.merge_cells("A1:B1")
    title_cell = ws_sum["A1"]
    title_cell.value = "MOSPL E2E LOAD TEST EXECUTIVE SUMMARY"
    title_cell.font = Font(bold=True, size=15, color="FFFFFF", name="Arial")
    title_cell.fill = _cell_fill("1A365D")  # Dark Slate Blue
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws_sum.row_dimensions[1].height = 35

    # Header Row (Row 3)
    for col in range(1, 3):
        cell = ws_sum.cell(row=3, column=col)
        cell.font = Font(bold=True, color="FFFFFF", size=11, name="Arial")
        cell.fill = _cell_fill("2B6CB0")  # Vibrant Slate Blue
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _thin_border()
    ws_sum.row_dimensions[3].height = 24

    # Data Rows (Rows 4-8)
    for row_idx, row in enumerate(ws_sum.iter_rows(min_row=4, max_row=8)):
        row_bg = "F7FAFC" if row_idx % 2 == 0 else "FFFFFF"
        for col_idx, cell in enumerate(row):
            cell.border = _thin_border()
            cell.font = Font(name="Arial", size=10)
            if col_idx == 0:
                cell.fill = _cell_fill("EDF2F7")  # Very Light Gray-Blue
                cell.font = Font(bold=True, name="Arial", size=10)
                cell.alignment = Alignment(horizontal="left", vertical="center")
            else:
                cell.fill = _cell_fill(row_bg)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                # Visual highlighting for Pass Percentage and Performance Score
                val_str = str(cell.value)
                if "%" in val_str:
                    cell.font = Font(bold=True, color="2F855A" if pass_percentage >= 90 else "C53030", name="Arial", size=10)
                elif "/ 100" in val_str:
                    cell.font = Font(bold=True, color="2F855A" if avg_perf_score >= 85 else "C53030", name="Arial", size=10)
        ws_sum.row_dimensions[4 + row_idx].height = 22

    ws_sum.column_dimensions["A"].width = 28
    ws_sum.column_dimensions["B"].width = 22

    # Add Chart
    # Write chart counts to clean cell range for referencing
    ws_sum["D15"] = "Status"
    ws_sum["E15"] = "Count"
    ws_sum["D16"] = "Passed"
    ws_sum["E16"] = passed
    ws_sum["D17"] = "Failed"
    ws_sum["E17"] = failed

    chart = PieChart()
    chart.title = "Load Test Pass vs Fail Status"
    chart.style = 10
    
    data = Reference(ws_sum, min_col=5, min_row=15, max_row=17)
    labels = Reference(ws_sum, min_col=4, min_row=16, max_row=17)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(labels)
    chart.dataLabels = DataLabelList()
    chart.dataLabels.showPercent = True
    chart.dataLabels.showVal = False
    
    ws_sum.add_chart(chart, "D3")

    # ─── Style "Load Test Details" ──────────────────────────────────────────
    ws_det = wb["Load Test Details"]
    ws_det.merge_cells("A1:F1")
    title_det = ws_det["A1"]
    title_det.value = "MOSPL E2E LOAD TEST DETAILED RESULTS"
    title_det.font = Font(bold=True, size=15, color="FFFFFF", name="Arial")
    title_det.fill = _cell_fill("1A365D")
    title_det.alignment = Alignment(horizontal="center", vertical="center")
    ws_det.row_dimensions[1].height = 35

    # Header Row (Row 3)
    for col in range(1, 7):
        cell = ws_det.cell(row=3, column=col)
        cell.font = Font(bold=True, color="FFFFFF", size=11, name="Arial")
        cell.fill = _cell_fill("2B6CB0")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _thin_border()
    ws_det.row_dimensions[3].height = 24

    # Data Rows
    status_styles = {
        "PASS": ("E6FFFA", "319795"),  # Soft Teal background, dark teal text
        "FAIL": ("FFF5F5", "E53E3E")   # Soft Red background, dark red text
    }

    for row_idx, row in enumerate(ws_det.iter_rows(min_row=4, max_row=total+3)):
        row_bg = "F7FAFC" if row_idx % 2 == 0 else "FFFFFF"
        status_val = str(row[5].value or "")
        
        bg_status, fg_status = status_styles.get(status_val, ("FFFFFF", "000000"))

        for col_idx, cell in enumerate(row):
            cell.border = _thin_border()
            cell.font = Font(name="Arial", size=9)
            
            if col_idx == 5:  # Status
                cell.fill = _cell_fill(bg_status)
                cell.font = Font(bold=True, color=fg_status, name="Arial", size=9)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_idx in [2, 3, 4]:  # Values, Thresholds, Results
                cell.fill = _cell_fill(row_bg)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.fill = _cell_fill(row_bg)
                cell.alignment = Alignment(horizontal="left", vertical="center")
        ws_det.row_dimensions[4 + row_idx].height = 20

    # Auto-adjust column widths
    column_widths = [15, 18, 16, 16, 18, 12]
    for i, w in enumerate(column_widths, 1):
        ws_det.column_dimensions[get_column_letter(i)].width = w

    ws_det.freeze_panes = "A4"
    ws_det.views.sheetView[0].showGridLines = True
    ws_sum.views.sheetView[0].showGridLines = True

    wb.save(filename)
    print(f"\n=================================================================")
    print(f"  [DONE] Load Test Report Generated:")
    print(f"  {filename}")
    print(f"  [STATS] Total: {total}  |  PASS: {passed}  |  FAIL: {failed}")
    print(f"  [RATE]  Pass Rate: {pass_percentage}%")
    print(f"=================================================================\n")
    return filename
