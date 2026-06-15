"""
appium_report_generator.py - Generates a rich Excel report for Appium Mobile test results.
Uses pandas + openpyxl for formatting, charts, and summary statistics.
"""
import os
import datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import PieChart, Reference
from openpyxl.chart.label import DataLabelList


OUTPUT_DIR = os.path.join(os.path.dirname(__file__), "..", "reports")


def _cell_fill(hex_color: str) -> PatternFill:
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")


def _thin_border() -> Border:
    s = Side(style="thin", color="BDBDBD")
    return Border(left=s, right=s, top=s, bottom=s)


def generate_appium_excel_report(test_results: list):
    """Generate a rich Excel report from the Appium mobile test results."""
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    timestamp = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    filename = os.path.join(OUTPUT_DIR, f"Mospl_Appium_Mobile_Report_{timestamp}.xlsx")

    df = pd.DataFrame(test_results)

    # Compute summary stats
    total = len(df)
    passed = (df["Status"] == "PASS").sum()
    failed = (df["Status"] == "FAIL").sum()
    skipped = (df["Status"] == "SKIP").sum()
    pass_rate = round((passed / total) * 100, 2) if total > 0 else 0

    category_summary = df.groupby("Category")["Status"].value_counts().unstack(fill_value=0)

    # Add ticks and emojis for the Excel display
    display_df = df.copy()
    display_df["Status"] = display_df["Status"].replace({
        "PASS": "PASS ✓",
        "FAIL": "FAIL ❌",
        "SKIP": "SKIP ⏭"
    })

    # Write sheets via pandas ExcelWriter
    with pd.ExcelWriter(filename, engine="openpyxl") as writer:
        # Sheet 1: All test cases
        display_df.to_excel(writer, sheet_name="All Test Cases", index=False, startrow=2)

        # Sheet 2: Summary
        summary_data = {
            "Metric": ["Total Tests", "Passed", "Failed", "Skipped", "Pass Rate (%)",
                        "Platform", "App Package", "Deployable Status"],
            "Value": [
                total, passed, failed, skipped, f"{pass_rate}%",
                "Android (UiAutomator2)", "com.mospl.mospl",
                "READY FOR DEPLOYMENT" if pass_rate >= 80 else "NOT READY - FIXES NEEDED"
            ],
        }
        pd.DataFrame(summary_data).to_excel(writer, sheet_name="Executive Summary", index=False, startrow=2)

        # Sheet 3: Category breakdown
        category_summary.to_excel(writer, sheet_name="Category Analysis")

        # Sheet 4: Failed tests only
        failed_df = display_df[display_df["Status"] == "FAIL ❌"]
        failed_df.to_excel(writer, sheet_name="Failed Tests", index=False, startrow=2)

        # Sheet 5: Passed tests only
        passed_df = display_df[display_df["Status"] == "PASS ✓"]
        passed_df.to_excel(writer, sheet_name="Passed Tests", index=False, startrow=2)

    # Apply rich formatting with openpyxl
    wb = load_workbook(filename)

    # Format "All Test Cases" sheet
    ws_all = wb["All Test Cases"]
    _add_title(ws_all, "MOSPL Mobile App -- Appium E2E Test Results", "0D47A1", cols=8)
    _format_header_row(ws_all, row=3, cols=8)
    _format_data_rows(ws_all, start_row=4, status_col=5)
    _set_col_widths(ws_all, [12, 16, 40, 55, 10, 14, 45, 22])
    ws_all.freeze_panes = "A4"

    # Format "Executive Summary" sheet
    ws_sum = wb["Executive Summary"]
    _add_title(ws_sum, "Executive Summary -- MOSPL Mobile Appium Testing", "0D47A1", cols=2)
    _format_header_row(ws_sum, row=3, cols=2)
    _format_summary_data(ws_sum, start_row=4, pass_rate=pass_rate)
    _set_col_widths(ws_sum, [30, 45])
    _add_pie_chart(ws_sum, passed, failed, skipped)

    # Format "Category Analysis" sheet
    ws_cat = wb["Category Analysis"]
    _format_category_sheet(ws_cat)

    # Format "Failed Tests" sheet
    if "Failed Tests" in wb.sheetnames:
        ws_fail = wb["Failed Tests"]
        _add_title(ws_fail, "Failed Tests -- Requires Attention", "B71C1C", cols=8)
        _format_header_row(ws_fail, row=3, cols=8, header_color="C62828")
        _format_data_rows(ws_fail, start_row=4, status_col=5)
        _set_col_widths(ws_fail, [12, 16, 40, 55, 10, 14, 45, 22])

    # Format "Passed Tests" sheet
    if "Passed Tests" in wb.sheetnames:
        ws_pass = wb["Passed Tests"]
        _add_title(ws_pass, "Passed Tests -- Verified & Stable", "1B5E20", cols=8)
        _format_header_row(ws_pass, row=3, cols=8, header_color="2E7D32")
        _format_data_rows(ws_pass, start_row=4, status_col=5)
        _set_col_widths(ws_pass, [12, 16, 40, 55, 10, 14, 45, 22])

    wb.save(filename)
    print(f"\n{'='*65}")
    print(f"  [DONE] Appium Mobile Excel Report Generated:")
    print(f"  {filename}")
    print(f"  [STATS] Total: {total}  |  PASS: {passed}  |  FAIL: {failed}")
    print(f"  [RATE]  Pass Rate: {pass_rate}%")
    print(f"{'='*65}\n")
    return filename


# ── Helper Functions ──────────────────────────────────────────────────────────

def _add_title(ws, title: str, bg_color: str, cols: int):
    ws.merge_cells(f"A1:{get_column_letter(cols)}1")
    cell = ws["A1"]
    cell.value = title
    cell.font = Font(bold=True, size=16, color="FFFFFF", name="Calibri")
    cell.fill = _cell_fill(bg_color)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30


def _format_header_row(ws, row: int, cols: int, header_color: str = "1565C0"):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
        cell.fill = _cell_fill(header_color)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _thin_border()
    ws.row_dimensions[row].height = 22


def _format_data_rows(ws, start_row: int, status_col: int):
    status_colors = {
        "PASS ✓": ("E8F5E9", "2E7D32"),
        "FAIL ❌": ("FFEBEE", "C62828"),
        "SKIP ⏭": ("FFF9C4", "F57F17"),
    }
    for row_idx, row in enumerate(ws.iter_rows(min_row=start_row)):
        status_cell = ws.cell(row=start_row + row_idx, column=status_col)
        status = str(status_cell.value or "")
        bg, fg = status_colors.get(status, ("FFFFFF", "000000"))
        row_bg = "E3F2FD" if row_idx % 2 == 0 else "FFFFFF"

        for cell in row:
            cell.border = _thin_border()
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            cell.font = Font(name="Calibri", size=10)
            if cell.column == status_col:
                cell.fill = _cell_fill(bg)
                cell.font = Font(bold=True, color=fg, size=10, name="Calibri")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.fill = _cell_fill(row_bg)

        ws.row_dimensions[start_row + row_idx].height = 18


def _format_summary_data(ws, start_row: int, pass_rate: float):
    for row in ws.iter_rows(min_row=start_row):
        label_cell = row[0]
        value_cell = row[1] if len(row) > 1 else None

        label_cell.font = Font(bold=True, size=11, name="Calibri")
        label_cell.alignment = Alignment(horizontal="left", vertical="center")
        label_cell.border = _thin_border()
        label_cell.fill = _cell_fill("E3F2FD")

        if value_cell:
            val = str(value_cell.value or "")
            value_cell.border = _thin_border()
            value_cell.alignment = Alignment(horizontal="center", vertical="center")
            value_cell.font = Font(size=11, name="Calibri")
            if "READY FOR DEPLOYMENT" in val:
                value_cell.fill = _cell_fill("C8E6C9")
                value_cell.font = Font(bold=True, color="1B5E20", size=12, name="Calibri")
            elif "NOT READY" in val:
                value_cell.fill = _cell_fill("FFCDD2")
                value_cell.font = Font(bold=True, color="B71C1C", size=12, name="Calibri")
            else:
                value_cell.fill = _cell_fill("FFFFFF")

        ws.row_dimensions[label_cell.row].height = 22


def _format_category_sheet(ws):
    for row in ws.iter_rows():
        for cell in row:
            cell.border = _thin_border()
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if cell.row == 1:
                cell.font = Font(bold=True, color="FFFFFF", name="Calibri")
                cell.fill = _cell_fill("1565C0")


def _set_col_widths(ws, widths: list):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _add_pie_chart(ws, passed: int, failed: int, skipped: int):
    """Add a pie chart to the Executive Summary sheet."""
    chart_data_start = 17
    ws[f"D{chart_data_start}"] = "Status"
    ws[f"E{chart_data_start}"] = "Count"
    ws[f"D{chart_data_start + 1}"] = "Passed"
    ws[f"E{chart_data_start + 1}"] = passed
    ws[f"D{chart_data_start + 2}"] = "Failed"
    ws[f"E{chart_data_start + 2}"] = failed
    ws[f"D{chart_data_start + 3}"] = "Skipped"
    ws[f"E{chart_data_start + 3}"] = skipped

    pie = PieChart()
    pie.title = "Mobile Test Results Overview"
    pie.style = 10

    data = Reference(ws, min_col=5, min_row=chart_data_start, max_row=chart_data_start + 3)
    labels = Reference(ws, min_col=4, min_row=chart_data_start + 1, max_row=chart_data_start + 3)
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    pie.dataLabels = DataLabelList()
    pie.dataLabels.showPercent = True
    pie.dataLabels.showCatName = True

    ws.add_chart(pie, "D3")
